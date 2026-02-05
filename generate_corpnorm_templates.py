
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------
DEEP_BLUE = "1F4B99"   # Header for core output
TEAL = "1ABC9C"        # Header for search columns
SOFT_YELLOW = "F5A623" # Header for combined column or special
LIGHT_GREY = "DDDDDD"  # Header background (title bar)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def setup_common_layout(ws, include_combined_column=False, title_note=None):
    """
    Set up:
    - Title bar with 'CorpNorm AI – By Kishor'
    - Headers
    - Column widths
    - Freeze panes
    """

    # 1) Title bar (A1:I2 or A1:J2 if combined column)
    last_col = "J" if include_combined_column else "I"
    title_range = f"A1:{last_col}2"
    ws.merge_cells(title_range)
    cell = ws["A1"]
    cell.value = "CorpNorm AI – By Kishor"
    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True, size=16)

    # 2) Optional title note in row 3
    if title_note:
        ws["A3"].value = title_note
        ws["A3"].font = Font(bold=True, size=11)
        ws.merge_cells(f"A3:{last_col}3")

    # 3) Header row (row 4)
    headers = [
        "Raw Company Name",
        "Normalized Company Name",
        "Website",
        "Industry",
        "Third Party Data Source Link",
        "Remark",
        "Search – Website",
        "Search – Industry",
        "Search – Profile / Registry"
    ]
    if include_combined_column:
        headers.append("All Searches")

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        # Decide header color
        if col_idx <= 6:
            fill_color = DEEP_BLUE
        elif col_idx <= 9:
            fill_color = TEAL
        else:
            fill_color = SOFT_YELLOW

        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(color="FFFFFF" if col_idx <= 9 else "000000", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 4) Column widths
    widths = {
        1: 40,  # Raw Company Name
        2: 40,  # Normalized Company Name
        3: 30,  # Website
        4: 25,  # Industry
        5: 40,  # Third Party Data Source Link
        6: 35,  # Remark
        7: 18,  # Search – Website
        8: 18,  # Search – Industry
        9: 22,  # Search – Profile / Registry
        10: 30  # All Searches (if present)
    }
    for col_idx, width in widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # 5) Freeze panes (keep title + header visible)
    # Data starts at row 5
    ws.freeze_panes = "A5"


def add_formulas_in_row(ws, row: int, include_combined_column=False):
    """
    Add search formulas in columns G, H, I (and J if needed) for the given row.
    """
    row_str = str(row)
    # G: Search – Website
    ws[f"G{row_str}"] = (
        '=HYPERLINK("https://www.google.com/search?q=" & '
        f'ENCODEURL($B{row_str} & " official website"), "Search Website")'
    )
    # H: Search – Industry
    ws[f"H{row_str}"] = (
        '=HYPERLINK("https://www.google.com/search?q=" & '
        f'ENCODEURL($B{row_str} & " industry"), "Search Industry")'
    )
    # I: Search – Profile / Registry
    ws[f"I{row_str}"] = (
        '=HYPERLINK("https://www.google.com/search?q=" & '
        f'ENCODEURL($B{row_str} & " company profile registry"), "Search Profile")'
    )

    if include_combined_column:
        # J: All Searches (combined)
        ws[f"J{row_str}"] = (
            '=HYPERLINK("https://www.google.com/search?q=" & '
            f'ENCODEURL($B{row_str} & " official website"), "Website")'
            ' & " | " & '
            'HYPERLINK("https://www.google.com/search?q=" & '
            f'ENCODEURL($B{row_str} & " industry"), "Industry")'
            ' & " | " & '
            'HYPERLINK("https://www.google.com/search?q=" & '
            f'ENCODEURL($B{row_str} & " company profile registry"), "Registry")'
        )


def add_usage_guide_sheet(wb, template_type: str):
    """
    Adds a 'Usage Guide' sheet to the workbook with instructions
    on how to use this specific template.
    template_type: one of "blank", "sample", "full"
    """
    ws_guide = wb.create_sheet(title="Usage Guide")

    lines = []

    # Common header
    lines.append("CorpNorm AI – By Kishor")
    lines.append("")
    lines.append("Usage Guide for Excel Template")
    lines.append("================================")
    lines.append("")

    if template_type == "blank":
        lines.append("Template Type: BLANK TEMPLATE")
        lines.append("")
        lines.append("Purpose:")
        lines.append("- This template is meant to be used with CorpNorm AI output.")
        lines.append("- It contains headers, formatting, and search formulas in row 5.")
    elif template_type == "sample":
        lines.append("Template Type: TEMPLATE + SAMPLE ROW")
        lines.append("")
        lines.append("Purpose:")
        lines.append("- This template shows one example row of CorpNorm AI output.")
        lines.append("- Use it to understand how data and search formulas work together.")
    elif template_type == "full":
        lines.append("Template Type: FULL WORKFLOW TEMPLATE")
        lines.append("")
        lines.append("Purpose:")
        lines.append("- This template is designed for the full CorpNorm AI workflow.")
        lines.append("- It includes combined search links and a short quick-start guide.")

    lines.append("")
    lines.append("Columns Overview (Main Sheet):")
    lines.append("- Column A: Raw Company Name")
    lines.append("- Column B: Normalized Company Name")
    lines.append("- Column C: Website")
    lines.append("- Column D: Industry")
    lines.append("- Column E: Third Party Data Source Link")
    lines.append("- Column F: Remark")
    lines.append("- Column G: Search – Website (Google search link)")
    lines.append("- Column H: Search – Industry (Google search link)")
    lines.append("- Column I: Search – Profile / Registry (Google search link)")
    lines.append("- Column J: All Searches (combined) – only in Full template")
    lines.append("")
    lines.append("How to Use with CorpNorm AI Output:")
    lines.append("1) Run the CorpNorm AI Streamlit app and download CorpNorm_Output.xlsx.")
    lines.append("2) Open this template in Excel.")
    lines.append("3) Copy data from CorpNorm_Output.xlsx and paste into columns A–F of the")
    lines.append("   main sheet (starting at row 5).")
    lines.append("4) Ensure the search formulas in row 5 (columns G–I, and J if present)")
    lines.append("   are filled correctly. Then drag/copy them down to all rows with data.")
    lines.append("5) Click the search links to quickly identify:")
    lines.append("   - Official websites")
    lines.append("   - Industry descriptions")
    lines.append("   - Third-party / registry profiles (LEI, Taiwantrade, D&B, etc.)")
    lines.append("")
    lines.append("Tips:")
    lines.append("- Always verify websites and profiles before finalizing your dataset.")
    lines.append("- Prefer official global corporate sites over distributors or social media.")
    lines.append("- If no official website exists, use a trusted registry or profile link.")
    lines.append("- Keep industries short and consistent (1–4 words, e.g., 'ELECTRONIC COMPONENTS').")
    lines.append("")
    lines.append("Branding:")
    lines.append("- The main sheet title is: 'CorpNorm AI – By Kishor'.")
    lines.append("- You can insert an official logo image above or near this title if you wish.")
    lines.append("")
    lines.append("Version:")
    lines.append("- CorpNorm Excel Template v1.0")

    # Write lines to the sheet
    row_idx = 1
    for line in lines:
        ws_guide.cell(row=row_idx, column=1, value=line)
        row_idx += 1

    # Optional: set column width
    ws_guide.column_dimensions["A"].width = 110


def create_blank_template(filename: str, include_combined_column=False):
    """
    Create a blank CorpNorm AI template:
    - Only headers and formulas (one row), no data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CorpNorm Template"

    title_note = "CorpNorm AI – Output Enrichment Template (Blank)"
    setup_common_layout(ws, include_combined_column=include_combined_column, title_note=title_note)

    # First data row is 5
    data_row = 5
    add_formulas_in_row(ws, data_row, include_combined_column=include_combined_column)

    # Add Usage Guide sheet
    add_usage_guide_sheet(wb, template_type="blank")

    wb.save(filename)
    print(f"Saved: {filename}")


def create_sample_template(filename: str, include_combined_column=False):
    """
    Create a template with one sample row filled in.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CorpNorm Template + Sample"

    title_note = "CorpNorm AI – Template with Sample Row (Example Only)"
    setup_common_layout(ws, include_combined_column=include_combined_column, title_note=title_note)

    data_row = 5

    # Sample row
    ws[f"A{data_row}"] = "Samsung Electro-Mechanics Co., Ltd."
    ws[f"B{data_row}"] = "SAMSUNG ELECTRO MECHANICS"
    ws[f"C{data_row}"] = "https://www.samsungsem.com/"
    ws[f"D{data_row}"] = "Electronic components"
    ws[f"E{data_row}"] = ""
    ws[f"F{data_row}"] = "Sample row – replace with your real data"

    # Style sample row lightly
    for col_idx in range(1, 11 if include_combined_column else 10):
        cell = ws.cell(row=data_row, column=col_idx)
        cell.border = thin_border

    add_formulas_in_row(ws, data_row, include_combined_column=include_combined_column)

    # Add Usage Guide sheet
    add_usage_guide_sheet(wb, template_type="sample")

    wb.save(filename)
    print(f"Saved: {filename}")


def create_full_workflow_template(filename: str):
    """
    Full workflow template:
    - Includes combined column (All Searches)
    - Includes title note with brief instructions
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CorpNorm Full Workflow"

    title_note = (
        "CorpNorm AI – Full Workflow: Paste your output in columns A–F. "
        "Use the search columns (G–J) to enrich Website / Industry / Profiles."
    )
    setup_common_layout(ws, include_combined_column=True, title_note=title_note)

    data_row = 5
    add_formulas_in_row(ws, data_row, include_combined_column=True)

    # Brief inline instructions on the main sheet
    guide_row = 6
    ws[f"A{guide_row}"] = (
        "Quick Instructions: 1) Paste CorpNorm_Output.xlsx data into columns A–F. "
        "2) Copy formulas in G–J down your used rows. "
        "3) Use the search links to find websites, industries, and registry profiles."
    )
    ws.merge_cells(f"A{guide_row}:J{guide_row}")
    ws[f"A{guide_row}"].alignment = Alignment(wrap_text=True)
    ws[f"A{guide_row}"].font = Font(size=10, italic=True)

    # Add Usage Guide sheet
    add_usage_guide_sheet(wb, template_type="full")

    wb.save(filename)
    print(f"Saved: {filename}")


if __name__ == "__main__":
    # 1) Blank Template (no sample row, no combined column)
    create_blank_template("CorpNorm_Template_Blank.xlsx", include_combined_column=False)

    # 2) Template + Sample Row (with three search columns)
    create_sample_template("CorpNorm_Template_Sample.xlsx", include_combined_column=False)

    # 3) Full Workflow Template (includes All Searches combined column)
    create_full_workflow_template("CorpNorm_Template_Full.xlsx")
